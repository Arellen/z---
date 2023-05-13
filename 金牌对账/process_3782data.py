import pandas as pd
from datetime import datetime

def process_3782data(id, erp_code, year, month, folder_path, file_path, sheet_name):
    # 定义查询条件
    previous_month = month - 1
    
    # 创建cost_items清单
    cost_items = ['线上引流', '三维家', '广宣品', '物流运费（捷鹏货运）', '物流运输费', '仓储费', '制服费', '驳回', '加急费用扣款', '运费', '安装费', '其他']

    # 读取Excel文件中的特定工作表
    df = pd.read_excel(file_path, sheet_name=sheet_name)

    # 读取额外的两个Excel文件
    detail_file_path = f"{folder_path}/{erp_code}费用明细导出.xlsx"
    balance_file_path = f"{folder_path}/{erp_code}费用余额导出.xlsx"
    detail_df = pd.read_excel(detail_file_path)
    balance_df = pd.read_excel(balance_file_path)

    # 列序列号
    id_column_index = 2
    erp_column_index = 6
    month_column_index = 7
    income_column_index = 8
    w_column_index = 22

    # 将月份列转换为日期类型，指定日期格式
    date_format = "%Y年%m月"  # 请根据实际的日期格式进行调整
    df.iloc[:, month_column_index] = pd.to_datetime(df.iloc[:, month_column_index], format=date_format, errors='coerce')

    # 从日期列中提取年份和月份
    df['year'] = df.iloc[:, month_column_index].apply(lambda x: x.year if pd.notnull(x) else -1)
    df['month'] = df.iloc[:, month_column_index].apply(lambda x: x.month if pd.notnull(x) else -1)

    # 执行查询
    filtered_df = df.loc[(df.iloc[:, id_column_index] == id) & (df.iloc[:, erp_column_index] == erp_code) & (df['year'] == year) & (df['month'] == previous_month)]
    
    # 获取满足条件的W列单元格的值
    values = filtered_df.iloc[:, w_column_index].sum()
    
    # 获取预收费用和费用扣款的O列的合计数
    prepayment_sum = detail_df.loc[(detail_df['加盟商|编码'] == id) & (detail_df['单据信息|业务类型'] == "预收费用"), '收入|金额'].sum()
    deduction_sum = detail_df.loc[(detail_df['加盟商|编码'] == id) & (detail_df['单据信息|业务类型'] == "费用扣款"), '收入|金额'].sum()

    # 获取cost_items里每一个item的值
    cost_items_values = {}
    for item in cost_items:
        cost_items_values[item] = detail_df.loc[(detail_df['加盟商|编码'] == id) & (detail_df['单据信息|业务类型']== "费用扣款") & (detail_df['备注'].str.contains(item)), '收入|金额'].sum()

    # 计算cost_items的总和
    cost_items_sum = sum(cost_items_values.values())

    # 如果deduction_sum 与 cost_items_sum 不相等，把cost_items中包含备注的值加上（deduction_sum - cost_items_sum）的值
    if deduction_sum != cost_items_sum:
        difference = deduction_sum - cost_items_sum
        cost_items_values['其他'] += difference

    # 获取费用余额导出表中的G列的值
    balance_value = balance_df.loc[balance_df['加盟商'] == id, '费用余额'].sum()

    # 四舍五入保留两位小数
    rounded_values = round(values, 2)
    rounded_prepayment_sum = round(prepayment_sum, 2)
    rounded_deduction_sum = round(deduction_sum, 2)
    rounded_balance_value = round(balance_value, 2)

    # 对cost_items每一个item的值进行四舍五入
    rounded_cost_items_values = {item: round(value, 2) for item, value in cost_items_values.items()}

    # 计算四舍五入后的cost_items值之和
    rounded_cost_items_sum = round(sum(rounded_cost_items_values.values()), 2)



    import openpyxl

    # 读取现有的 Excel 文件
    workbook = openpyxl.load_workbook(file_path)

    # 获取要编辑的工作表
    worksheet = workbook[sheet_name]

    # 用于存储满足条件的行号
    target_row = -1

    # 从第二行开始遍历，因为通常第一行是标题行
    for row in range(2, worksheet.max_row + 1):
        # 检查 id_column_index 列的值
        cell_id = worksheet.cell(row=row, column=id_column_index + 1).value
        
        # 检查 erp_column_index 列的值
        cell_erp_code = worksheet.cell(row=row, column=erp_column_index + 1).value
        
        # 检查 month_column_index 列的值，并转换为字符串
        cell_date = worksheet.cell(row=row, column=month_column_index + 1).value
        
        if isinstance(cell_date, datetime):
            cell_date_str = cell_date.strftime("%Y-%m-%d")
        else:
            cell_date_str = str(cell_date)
        
        # 检查所有条件是否满足
        if cell_id == id and cell_erp_code == erp_code and cell_date_str == f"{year:02}-{month:02}-01":
            target_row = row
            break 

        # 比较四舍五入后的值
    if round(rounded_values + rounded_prepayment_sum + rounded_deduction_sum, 2) == rounded_balance_value and rounded_deduction_sum == rounded_cost_items_sum:
            
            # 将rounded_prepayment_sum值写入第11列
            worksheet.cell(row=target_row, column=10, value=rounded_prepayment_sum)

            # 将rounded_cost_items_values的值依次写入第12列，第13列一直到第23列
            for idx, value in enumerate(rounded_cost_items_values.values()):
                worksheet.cell(row=target_row, column=11 + idx, value=-value)
            # 将rounded_prepayment_sum值写入第11列
            worksheet.cell(row=target_row, column=22, value=rounded_balance_value)
            # 保存修改后的工作表
            workbook.save(file_path)

    else:
        print(f"id: {id}"f"---{previous_month}月份余额+: {rounded_values}+"f"当月收入+: {rounded_prepayment_sum}+"f"当月支出+: {rounded_deduction_sum}="f"当月余额=: {rounded_balance_value}")
        # for item, value in rounded_cost_items_values.items():
        #     print(f"{item}: {value}")
